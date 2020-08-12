"""
  Implementation of task 4 from module four
"""

import openpyxl

XLSX_FILE_NAME = 'sheet.xlsx'
BEFORE_INVERT_SHEET_NAME = 'before'
AFTER_INVERT_SHEET_NAME = 'after'


def invert_spreadsheet():
    """
    Method responsible for getting values from  before sheet and transposing them in after sheer
    """

    wb = openpyxl.load_workbook(XLSX_FILE_NAME)

    if AFTER_INVERT_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[AFTER_INVERT_SHEET_NAME])

    before_sheet = wb[BEFORE_INVERT_SHEET_NAME]
    wb.create_sheet(AFTER_INVERT_SHEET_NAME)
    after_sheet = wb[AFTER_INVERT_SHEET_NAME]
    for column_index in range(1, before_sheet.max_column + 1):
        for row_index in range(1, before_sheet.max_row + 1):
            after_sheet.cell(row=column_index, column=row_index).value = \
                before_sheet.cell(row=row_index, column=column_index).value
    wb.save(XLSX_FILE_NAME)


if __name__ == "__main__":
    invert_spreadsheet()
